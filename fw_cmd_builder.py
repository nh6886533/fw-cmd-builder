#!/usr/bin/python3

#public module
import os,re
from openpyxl import load_workbook
from IPy import IP

class InputFile:
	'''
	读取input.xlsx文件,文件格式如下：
	A2：SRC_ZONE,	B2：SRC_ADD,		C2：SRC_GROUP
	D2：DST_ZONE,	E2：DST_ZONE,	F2：DST_GROUP
	G2：TCP OR UDP	H3：PORT			I2：PORT_GROUP
	J2：ACTION
	sheet名称为厂商识别代码：如hv5
	'''
	def __init__(self, path):
		self.wb = load_workbook(path)
		self.ws = self.wb.worksheets[0]	

	def get_vender(self):
		'''get vender information (str)
		'''
		return self.wb.sheetnames[0]

	def get_address(self, who):
		'''
		return a list contain address information{'add','group','zone'}
		who MUST be src or dst
		'''
		res = []
		row = 3
		column = {'src':1, 'dst':4}.get(who)
		while True:				
			if self.ws.cell(row = row, column = column).value == None: break			
			one_add = {}
			one_add['zone'] = str(self.ws.cell(row = row, column = column).value).strip()
			one_add['add'] = str(self.ws.cell(row = row, column = column+1).value).strip()
			one_add['group'] = str(self.ws.cell(row = row, column = column+2).value).strip()
			res.append(one_add)
			row += 1
		return res

	def get_service(self):
		'''return a list contain service information{'proto','port','group'}
		'''
		res = []
		row,column = 3,7
		while True:				
			if self.ws.cell(row = row, column = column).value == None: break
			one_port = {}
			one_port['proto'] = str(self.ws.cell(row = row, column = column).value).strip()
			one_port['port'] = str(self.ws.cell(row = row, column = column+1).value).strip()
			one_port['group'] = str(self.ws.cell(row = row, column = column+2).value).strip()
			res.append(one_port)
			row += 1
		return res

	def get_action(self):
		'''return policy action which shoule be permit or deny
		'''
		return self.ws.cell(row = 3, column = 10).value

class FwCmdBuilder:
	def __init__(self, src_add, dst_add, service, action):
		'''
		zone,add,add_group,service dict
		'''
		#src_add & dst_add format [{'zone':xxx, 'add':xxx, 'mask':xxx, group':xxx}...]
		self.src_add = self.get_address(src_add)
		self.dst_add = self.get_address(dst_add)
		#servic format [{'proto':xxx,'port':xxx,'group':xxx,'range':True/False}]
		self.service = self.get_service(service)
		self.action = action

	def get_address(self, address):
		'''
		对地址进行格式整理
		address is a list contain {'add','group','zone'}
		'''
		for i in address:
			#range
			if '-' in i['add']:
				net = re.match('(\S+)\.\d+-', i['add']).group(1)
				range_start = re.search('\.(\d+)-', i['add']).group(1)
				range_end = re.search('-(\d+)', i['add']).group(1)
				i['add'] = net+'.'+range_start+' '+net+'.'+range_end
				i['mask'] = 'range'
			#ip/mask
			elif '/' in i['add']:
				i['mask'] = re.search('(/\d+)',i['add']).group(1)
				i['add'] = re.match('(\S+)/', i['add']).group(1)
			#any
			elif re.match('[Aa][Nn][Yy]', i['add']):
				i['add'] = 'any'
				i['mask'] = ''
				i['group'] = 'any'
			#ip
			else:
				i['mask'] = '/32'
		return address

	def get_service(self, service):
		'''
		'''
		for i in service:
			#range
			if '-' in str(i['port']):
				i['range'] = True
				i['port'] = str(i['port']).replace('-',' ')
			#any
			elif re.match('[Aa][Nn][Yy]', str(i['port'])):
				i['port'] = 'any'
				i['group'] = 'any'
				i['range'] = False
			else:
				i['range'] = False				
		return service

class HillStoneBuilderV5(FwCmdBuilder):
	def __init__(self, src_add, dst_add, service, action):
		super().__init__(src_add, dst_add, service, action)

	def make_ip(self, who):
		'''
		对没有定义地址组的孤立地址编写配置
		形参：who Must be src or dst
		'''
		res = ''
		for i in {'src':self.src_add, 'dst':self.dst_add}.get(who):
			if i['group'] == 'None':
				if i['add'] == 'any':
					res+={'src':'src-addr', 'dst':'dst-addr'}.get(who)+' Any\n'
				else:
					res+={'src':'src-ip', 'dst':'dst-ip'}.get(who)+' '+i['add']+i['mask']+'\n'
		return res

	def make_address(self, group):
		'''
		对地址组编写配置
		'''
		res = 'address ' + group+'\n'
		for i in self.src_add+self.dst_add:
			if i['group'] == group:
				if i['mask'] == 'range':
					res += 'range '+i['add']+'\n'
				else:
					res += 'ip '+i['add']+i['mask']+'\n'
		res+= 'exit\n'
		return res

	def make_service(self, group):
		'''
		编写服务组
		'''
		res = 'service '+group+'\n'
		for i in self.service:
			if i['group'] == group:
				res += ' '.join([i['proto'],'dst-port',i['port']])
				res +='\n'
		res+= 'exit\n'
		return res

	def make_policy(self, action, src_groups, dst_groups, ser_groups):
		'''
		编写policy
		'''
		res = 'rule\n'
		res += 'action '+action+'\n'
		res += 'src-zone '+self.src_add[0]['zone']+'\n'
		res += 'dst-zone '+self.dst_add[0]['zone']+'\n'
		#src ip and address
		res += self.make_ip('src')
		for i in src_groups: res += 'src-addr '+i+'\n'
		#dst ip and address
		res += self.make_ip('dst')		
		for i in dst_groups: res += 'dst-addr '+i+'\n'
		#service
		for i in ser_groups: res += 'service '+i+'\n'
		res += 'exit\n'
		return res

class HillStoneBuilderV4(FwCmdBuilder):
	def __init__(self, src_add, dst_add, service, action):
		super().__init__(src_add, dst_add, service, action)

	def make_address(self, who, group):
		'''
		对地址组编写配置
		形参：who MUST be src or dst
		'''
		addr = {'src':self.src_add, 'dst':self.dst_add}.get(who)
		res = 'address '+group+'\n'
		res += 'reference-zone '+addr[0]['zone']+'\n'
		for i in addr:
			if i['group'] == group:
				if i['mask'] == 'range':
					res += 'range '+i['add']+'\n'
				else:
					res += 'ip '+i['add']+i['mask']+'\n'
		res+= 'exit\n'
		return res	

	def make_service(self, group):
		'''
		编写服务组
		'''
		res = 'service '+group+'\n'
		for i in self.service:
			if i['group'] == group:
				res += ' '.join([i['proto'],'dst-port',i['port']])
				res +='\n'
		res+= 'exit\n'
		return res

	def make_policy(self, action, src_groups, dst_groups, ser_groups):
		'''
		编写policy
		'''
		res = "policy from "+self.src_add[0]['zone']+" to "+self.dst_add[0]['zone']+'\n'
		res += 'rule\n'
		res += 'action '+action+'\n'
		#src ip and address
		for i in src_groups: res += 'src-addr '+i+'\n'
		#dst ip and address	
		for i in dst_groups: res += 'dst-addr '+i+'\n'
		#service
		for i in ser_groups: res += 'service '+i+'\n'
		res += 'exit\n'	
		return res		

class CiscoBuilder(FwCmdBuilder):
	def __init__(self, src_add, dst_add, service, action):
		super().__init__(src_add, dst_add, service, action)

	def make_address(self, group):
		'''
		对地址组编写配置
		'''
		res = 'object-group network ' + group+'\n'
		for i in self.src_add+self.dst_add:
			if i['group'] == group:
				#对range地址按每个host编写
				if i['mask'] == 'range':
					[range_start, range_end] = i['add'].split(' ')
					range_start = range_start.split('.')
					range_end = range_end.split('.')
					for host in range(int(range_start[3]),int(range_end[3])+1):
						res += 'network-object host '+'.'.join(range_start[0:3])+'.'+str(host)+'\n'
				#对host地址进行编写
				elif i['mask'] == '/32':
					res += 'network-object host '+i['add']+'\n'
				#对net地址进行编写
				else:
					#将1.1.1.0/24转换成1.1.1.0/255.255.255.0
					add_mask = IP(i['add']+i['mask']).strNormal(2)
					res += 'network-object '+' '.join(add_mask.split('/'))+'\n'
		res += 'exit\n'			
		return res

	def make_service(self, group):
		'''
		编写服务组
		''' 
		res = ''
		for i in self.service:
			if i['group'] == group:
				proto = i['proto']
				if i['range']:
					res += 'port-object range '+i['port']+'\n'
				else:
					res += 'port-object eq '+i['port']+'\n'
		res = 'object-group service '+group+' '+proto+'\n'+res
		res += 'exit\n'
		return res

	def make_policy(self, acl_name, action, src, dst, proto, ser):
		'''
		编写ACL
		'''
		return ' '.join(['access-list',acl_name,'extended',action,proto,src,dst,ser])

class FortinetBuilder(FwCmdBuilder):
	def __init__(self):
		pass

class NetscreenBuilder(FwCmdBuilder):
	def __init__(self, src_add, dst_add, service, action):
		super().__init__(src_add, dst_add, service, action)

	def make_address(self, addr):
		'''
		编写地址
		形参：addr 为self.src_add或self.dst_add的元素
		返回一条地址编写配置
		'''
		#range模式编写
		if addr['mask'] == 'range':
			res = ''
			[range_start, range_end] = addr['add'].split(' ')
			range_start = range_start.split('.')
			range_end = range_end.split('.')			
			for host in range(int(range_start[3]),int(range_end[3])+1):
				res += 'set address '+addr['zone']+' '+'.'.join(range_start[0:3])+'.'+str(host)+'/32 '
				res += '.'.join(range_start[0:3])+'.'+str(host)+' 255.255.255.255\n'
		#孤立地址编写
		else:
			res = ['set address',addr['zone'],addr['add']+addr['mask']]
			ip_mask = IP(addr['add']+addr['mask'])
			ip_mask.NoPrefixForSingleIp = None
			ip_mask = ip_mask.strNormal(2).replace('/',' ')
			res.append(ip_mask)
			res = ' '.join(res)+'\n'
		return res

	def make_group_address(self, group):
		'''
		对地址组编写配置
		'''
		res = ''
		for i in self.src_add+self.dst_add:
			if i['group'] == group:
				#对range地址按每个host编写
				if i['mask'] == 'range':
					[range_start, range_end] = i['add'].split(' ')
					range_start = range_start.split('.')
					range_end = range_end.split('.')
					for host in range(int(range_start[3]),int(range_end[3])+1):
						res += ' '.join(['set group address',i['zone'],group,'add','.'.join(range_start[0:3])+'.'+str(host)+'/32'])+'\n'
				else:
					res += ' '.join(['set group address',i['zone'],group,'add',i['add']+i['mask']])+'\n'
		res = 'set group address '+group+'\n'+res
		return res

	def make_service(self, ser):
		'''
		编写service
		'''
		res = 'set service '+ser['proto'].upper()+'_'+ser['port'].replace(' ','-')+' protocol '
		res += ser['proto'].lower()+' src-port 0-65535 dst-port '+ser['port']+'-'+ser['port']+'\n'
		return res

	def make_group_service(self, group):
		'''
		编写地址组策略
		'''	
		res = 'set group service '+group+'\n'
		for i in self.service:
			if i['group'] == group:
				res += 'set group service '+group+' add '+i['proto'].upper()+'_'+i['port'].replace(' ','-')+'\n'
		return res

	def make_policy(self, src, dst, service, action):
		'''
		策略配置编写
		'''
		res = 'set policy from '+self.src_add[0]['zone']+' to '+self.dst_add[0]['zone']+' '
		res += ' '.join([src,dst,service,action])
		return res

class vender_manager:
	'''对多厂商设备进行识别，调用不同的防火墙配置生成脚本
	目前支持厂商：
	1.山石v5版本
	2.山石v4版本
	3.思科ASA
	4.Juniper Netscreen
	'''
	def __init__(self, path):
		'''读取input.xlsx文件
		并按照识别代码调用对应的防火墙配置编写代码
		'''
		self.path = path
		file = InputFile(os.path.join(self.path, 'input.xlsx'))
		policy_info = {
		'src_add':file.get_address('src'),
		'dst_add':file.get_address('dst'),
		'service':file.get_service(),
		'action':file.get_action()
		}		
		{'hv5':self.hillstone_v5,
		'hv4':self.hillstone_v4,
		'c':self.cisco,
		'n':self.netscreen}.get(file.get_vender())(**policy_info)

	def hillstone_v5(self, src_add, dst_add, service, action):
		'''山石防火墙v5版本策略生成管理函数
		'''
		builder = HillStoneBuilderV5(src_add, dst_add, service, action)
		res = 'hillstone v5 command'.center(50, '=')+'\n'
		#address group commands
		src_groups = set([i['group'] for i in builder.src_add if i['group'] != 'None'])
		dst_groups = set([i['group'] for i in builder.dst_add if i['group'] != 'None'])
		for i in list(src_groups)+list(dst_groups):
			if i != 'any':
				res += builder.make_address(i)
				res += '\n'
		#service group commands
		ser_groups = set([i['group'] for i in builder.service if i['group'] != 'None'])
		for i in ser_groups:
			if i != 'any':
				res += builder.make_service(i)
				res += '\n'
		#policy commands
		res += builder.make_policy(action, src_groups, dst_groups, ser_groups)
		self.write_output(res)

	def hillstone_v4(self, src_add, dst_add, service, action):
		'''山石防火墙v4版本策略生成管理函数
		'''
		builder = HillStoneBuilderV4(src_add, dst_add, service, action)
		res = 'hillstone v4 command'.center(50, '=')+'\n'
		#src address group commands
		src_groups = set([i['group'] for i in builder.src_add if i['group'] != 'None'])
		for i in src_groups:
			if i != 'any':
				res += builder.make_address('src', i)
				res += '\n'
		#dst address group commands	
		dst_groups = set([i['group'] for i in builder.dst_add if i['group'] != 'None'])
		for i in dst_groups:
			if i != 'any':			
				res += builder.make_address('dst', i)
				res += '\n'
		#service group commands
		ser_groups = set([i['group'] for i in builder.service if i['group'] != 'None'])
		for i in ser_groups:
			if i != 'any':
				res += builder.make_service(i)
				res += '\n'
		#policy commands
		res += builder.make_policy(action, src_groups, dst_groups, ser_groups)
		self.write_output(res)

	def cisco(self, src_add, dst_add, service, action):
		'''思科ASA策略生成管理函数
		'''
		builder = CiscoBuilder(src_add, dst_add, service, action)
		res = 'cisco ASA command'.center(50, '=')+'\n'
		#address group commands
		src_groups = set([i['group'] for i in builder.src_add if i['group'] != 'None'])
		dst_groups = set([i['group'] for i in builder.dst_add if i['group'] != 'None'])
		for i in list(src_groups)+list(dst_groups):
			if i != 'any':
				res += builder.make_address(i)
				res += '\n'
		#service group commands
		ser_groups = set([i['group'] for i in builder.service if i['group'] != 'None'])
		for i in ser_groups:
			if i != 'any':
				res += builder.make_service(i)
				res += '\n'
		#判断是否需要定义源地址组
		if len(src_groups) == 0:
			src = builder.src_add[0]
			src = IP(src['add']+src['mask'])
			src.NoPrefixForSingleIp = None
			src = src.strNormal(2).replace('/',' ')
		else:
			src = src_groups.pop()
			if src != 'any': src = 'object-group '+src
		#判断是否需要定义目标地址组
		if len(dst_groups) == 0:
			dst = builder.dst_add[0]
			dst = IP(dst['add']+dst['mask'])
			dst.NoPrefixForSingleIp = None
			dst = dst.strNormal(2).replace('/',' ')
		else:
			dst = dst_groups.pop()
			if dst != 'any': dst = 'object-group '+dst
		#判断是否需要定义服务组
		if len(ser_groups) == 0:
			service = builder.service[0]
			if service['range']:
				ser = 'range '+service['port']
			else:
				ser = 'eq '+service['port']
		else:
			ser = ser_groups.pop()
			if ser == 'any':
				ser = ''
			else:
				ser = 'object-group '+ser
		#policy command
		res += builder.make_policy(builder.src_add[0]['zone'], action, src, dst, builder.service[0]['proto'], ser)
		self.write_output(res)

	def netscreen(self, src_add, dst_add, service, action):
		'''netscreen策略生成管理函数
		'''
		builder = NetscreenBuilder(src_add, dst_add, service, action)
		res = 'Juniper Netscreen command'.center(50, '=')+'\n'
		#src address command
		for i in builder.src_add:
			if i['add'] != 'any':
				res += builder.make_address(i)
		res += '\n'
		#src address group command
		src_groups = list(set([i['group'] for i in builder.src_add if i['group'] not in ['None','any']]))
		for i in src_groups:
			if i != 'any':
				res += builder.make_group_address(i)
		res += '\n'
		#dst address command
		for i in builder.dst_add:
			if i['add'] != 'any':
				res += builder.make_address(i)
		res += '\n'
		#dst address group command
		dst_groups = list(set([i['group'] for i in builder.dst_add if i['group'] not in ['None','any']]))
		for i in dst_groups:
			if i != 'any':
				res += builder.make_group_address(i)
		res += '\n'
		#servie command
		for i in builder.service:
			if i['port'] != 'any':
				res += builder.make_service(i)
		res += '\n'
		#service group command
		ser_groups = list(set([i['group'] for i in builder.service if i['group'] not in ['None','any']]))
		for i in ser_groups:
			res += builder.make_group_service(i)
		res += '\n'
		#policy command
		#判断是否有源地址组
		if len(src_groups) == 0:
			src = builder.src_add[0]
			if src['add'] != 'any': 
				src = src['add']+src['mask']
			else:
				src = 'ANY'
		else:
			src = src_groups[0]
		#判断是否有目标地址组
		if len(dst_groups) == 0:
			dst = builder.dst_add[0]
			if dst['add'] != 'any':
				dst = dst['add']+dst['mask']
			else:
				dst = 'ANY'
		else:
			dst = dst_groups[0]
		#判断是否有服务地址组
		if len(ser_groups) == 0:
			ser = builder.service[0]
			if ser['port'] != 'any':
				ser = ser['proto'].upper()+'_'+ser['port']
			else:
				ser = 'ANY'
		else:
			ser = ser_groups[0]
		res += builder.make_policy(src,dst,ser,action)
		self.write_output(res)

	def write_output(self, data):
		'''
		return command to ouput.txt
		'''
		with open(os.path.join(self.path, 'output.txt'), 'w') as file:
			file.write(data)

def main():
	vender_manager(os.getcwd())

if __name__ == '__main__': 
	main()

